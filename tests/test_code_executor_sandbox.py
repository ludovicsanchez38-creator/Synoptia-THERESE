"""US-001 : durcissement du sandbox d'exécution de code (skills Office).

Deux protections :
1. Blocage de l'introspection par dunders (évasion classique du namespace
   restreint : ().__class__.__bases__[0].__subclasses__()).
2. Exécution dans un sous-process spawn isolé : l'interpréteur enfant n'hérite
   pas de la mémoire du backend (token de session, clé Fernet), donc une
   éventuelle évasion ne donne pas accès aux secrets du process principal.
"""
import queue

import pytest
from app.services.skills.code_executor import (
    CodeExecutionError,
    _run_generation_in_subprocess,
    execute_sandboxed,
    validate_code,
)

XLSX_OK = "wb = Workbook()\nws = wb.active\nws['A1'] = title\nwb.save(output_path)"


@pytest.mark.parametrize(
    "snippet",
    [
        "x = ().__class__.__bases__",
        "x = ().__class__.__subclasses__()",
        "x = type.__mro__",
        "x = (1).__class__.__globals__",
        "x = ().__class__.__base__.__subclasses__()",
    ],
)
def test_validate_code_bloque_introspection_dunder(snippet):
    is_valid, _ = validate_code(snippet)
    assert is_valid is False


def test_validate_code_accepte_du_code_office_legitime():
    is_valid, msg = validate_code(XLSX_OK)
    assert is_valid is True, msg


def test_worker_genere_le_fichier_et_signale_ok(tmp_path):
    out = tmp_path / "test.xlsx"
    rq: queue.Queue = queue.Queue()
    _run_generation_in_subprocess(XLSX_OK, str(out), "Mon titre", "xlsx", 10, rq)
    status, _ = rq.get_nowait()
    assert status == "ok"
    assert out.exists()


def test_worker_signale_erreur_sur_code_qui_leve(tmp_path):
    rq: queue.Queue = queue.Queue()
    _run_generation_in_subprocess(
        "raise ValueError('boum')", str(tmp_path / "x.xlsx"), "T", "xlsx", 10, rq
    )
    status, detail = rq.get_nowait()
    assert status == "error"
    assert "boum" in detail


@pytest.mark.asyncio
async def test_execute_sandboxed_rejette_evasion_dunder(tmp_path):
    code = "wb = Workbook()\nx = ().__class__.__bases__\nwb.save(output_path)"
    with pytest.raises(CodeExecutionError):
        await execute_sandboxed(code, str(tmp_path / "x.xlsx"), "T", "xlsx")


@pytest.mark.asyncio
async def test_execute_sandboxed_passe_par_un_sous_process_spawn(monkeypatch, tmp_path):
    """L'exécution doit transiter par un Process spawn isolé, pas par le thread
    du backend (sinon une évasion accéderait aux secrets en mémoire).

    B-048 : ce test ne comparait que le `__name__` de la cible capturée, et le
    faux `Process` jetait `args` et `daemon`. Un leurre portant le même nom et
    lisant les secrets satisfaisait l'assertion. On compare désormais
    l'IDENTITÉ de la fonction, et on conserve les arguments et le drapeau
    `daemon` pour les éprouver.

    Il éprouve la FORME de l'appel, ce qui reste utile mais ne prouve aucun
    isolement : la preuve d'isolement est le test qui suit, qui lance un vrai
    sous-processus.
    """
    import multiprocessing as mp

    captured: dict = {}
    real_get_context = mp.get_context

    class _Q:
        def get(self, timeout=None):
            return ("ok", "")

    class _P:
        def __init__(self, target, args, daemon=False):
            captured["target"] = target
            captured["args"] = args
            captured["daemon"] = daemon

        def start(self):
            pass

        def is_alive(self):
            return False

        def join(self, t=None):
            pass

        def terminate(self):
            pass

    class _Ctx:
        def Queue(self):
            return _Q()

        def Process(self, target, args, daemon=False):
            return _P(target, args, daemon)

    monkeypatch.setattr(
        mp, "get_context", lambda m: _Ctx() if m == "spawn" else real_get_context(m)
    )

    sortie = str(tmp_path / "x.xlsx")
    await execute_sandboxed(XLSX_OK, sortie, "T", "xlsx", nb_slides=7)

    assert captured["target"] is _run_generation_in_subprocess, (
        "un leurre portant le même nom ne doit pas satisfaire ce test"
    )
    assert captured["daemon"] is True, (
        "le sous-processus doit être daemon : sinon un backend qui s'arrête "
        "attend un enfant échappé au lieu de le tuer"
    )
    code, chemin, titre, format_type, nb_slides, _queue = captured["args"]
    assert (code, chemin, titre, format_type, nb_slides) == (
        XLSX_OK,
        sortie,
        "T",
        "xlsx",
        7,
    )


@pytest.mark.asyncio
async def test_le_sous_processus_n_herite_pas_de_la_memoire_du_backend(
    monkeypatch, tmp_path
):
    """La promesse en tête de fichier, éprouvée sur un VRAI sous-processus.

    B-048 : aucun test du fichier n'observait l'enfant. Les quatre autres
    appels à `_run_generation_in_subprocess` l'appellent comme une fonction
    ordinaire - donc dans le processus de test - et le seul qui parlait de
    spawn le remplaçait par un faux qui n'exécute rien.

    Le témoin est posé dans la MÉMOIRE du parent (on détourne
    `_build_namespace` pour injecter un titre témoin). Un interpréteur `spawn`
    réimporte le module à neuf : le détournement n'existe pas chez lui, et le
    fichier produit porte le vrai titre. Une exécution en processus (appel
    direct, ou contexte `fork`) porterait le témoin.

    Le témoin transite par le FICHIER produit, pas par une variable du parent :
    il reste donc observable même si le sabotage est un `fork`, où une liste
    partagée serait, elle, invisible.
    """
    import app.services.skills.code_executor as ce
    from openpyxl import load_workbook

    vrai_build_namespace = ce._build_namespace

    def _build_namespace_temoin(output_path, title, format_type, nb_slides=10):
        namespace = vrai_build_namespace(output_path, title, format_type, nb_slides)
        namespace["title"] = "TEMOIN-MEMOIRE-DU-PARENT"
        return namespace

    monkeypatch.setattr(ce, "_build_namespace", _build_namespace_temoin)

    sortie = tmp_path / "isolement.xlsx"
    await execute_sandboxed(XLSX_OK, str(sortie), "TITRE-ATTENDU", "xlsx")

    assert sortie.exists(), "le sous-processus n'a rien produit"
    valeur = load_workbook(str(sortie)).active["A1"].value
    assert valeur == "TITRE-ATTENDU", (
        "le code a été exécuté dans un interpréteur qui voyait la mémoire du "
        f"parent (A1 = {valeur!r})"
    )


def test_pandas_est_un_import_interdit():
    """Passe 4 : pandas.read_csv / to_excel ne passent pas par open()
    et atteignaient ~/.therese, ~/.ssh, ou une URL HTTP. Le prompt du
    skill ne le proposait pas ; le validateur l'acceptait quand même."""
    from app.services.skills.code_executor import _validate_imports

    ok, msg = _validate_imports("import pandas as pd\npd.read_csv('/etc/passwd')", "xlsx")
    assert ok is False
    assert "pandas" in msg.lower()


def test_save_via_variable_ecrit_dans_le_dossier_de_sortie(tmp_path):
    """wb.save(chemin) avec une variable passait au travers de la
    réécriture, qui ne couvrait qu'un littéral. _ensure_save_call
    s'arrêtait dès qu'un .save( existait."""
    dehors = tmp_path / "dehors"
    dedans = tmp_path / "dedans"
    dehors.mkdir()
    dedans.mkdir()
    hors_cible = dehors / "evasion.xlsx"
    sortie = dedans / "autorise.xlsx"
    code = (
        "wb = Workbook()\n"
        "ws = wb.active\n"
        "ws['A1'] = title\n"
        f"chemin = {str(hors_cible)!r}\n"
        "wb.save(chemin)\n"
    )
    rq: queue.Queue = queue.Queue()
    _run_generation_in_subprocess(code, str(sortie), "T", "xlsx", 10, rq)
    status, detail = rq.get_nowait()
    assert status == "ok", detail
    assert sortie.exists(), "le fichier attendu n'a pas été écrit"
    assert not hors_cible.exists(), (
        "wb.save(chemin) a écrit hors du dossier de sortie"
    )


def test_une_bibliotheque_ne_lit_pas_hors_du_dossier(tmp_path):
    """open() du namespace est absent ; openpyxl.load_workbook, lui,
    utilise le vrai open du process. C'est le trou : le code généré
    lit ~/.therese via la bibliothèque, pas via open().

    Le secret est DANS UN AUTRE dossier que la sortie : le même
    tmp_path les laisserait passer (le garde autorise le dossier
    de sortie, pas le disque entier).
    """
    from openpyxl import Workbook as _WB

    dehors = tmp_path / "dehors"
    dedans = tmp_path / "dedans"
    dehors.mkdir()
    dedans.mkdir()
    secret = dehors / "secret.xlsx"
    wb = _WB()
    wb.active["A1"] = "mot de passe imap"
    wb.save(secret)

    sortie = dedans / "autorise.xlsx"
    code = (
        "from openpyxl import load_workbook\n"
        f"wb = load_workbook({str(secret)!r})\n"
        "wb.save(output_path)\n"
    )
    rq: queue.Queue = queue.Queue()
    _run_generation_in_subprocess(code, str(sortie), "T", "xlsx", 10, rq)
    status, detail = rq.get_nowait()
    assert status == "error", f"la bibliothèque a lu hors cible : {detail}"
    assert "PermissionError" in detail or "hors" in detail.lower()
    assert not sortie.exists()


def test_la_garde_reseau_refuse_une_connexion():
    """pandas.read_csv('https://…') parlait HTTP via urllib, pas open().
    La garde s'installe dans le sous-processus avant exec."""
    import socket

    import app.services.skills.code_executor as ce

    assert hasattr(ce, "_installer_garde_reseau"), "la garde réseau n'existe pas"
    restaurer = ce._installer_garde_reseau()
    try:
        with pytest.raises(PermissionError):
            socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    finally:
        restaurer()

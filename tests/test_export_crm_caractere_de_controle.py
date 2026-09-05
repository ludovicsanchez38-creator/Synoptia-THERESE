"""Une seule fiche empoisonnee ne doit pas emporter tout l'export CRM.

02/09/2026, campagne de robustesse du cycle 2 (RB-006, B-163). Un contact est
enregistre avec un caractere de controle dans sa societe (`b\\x07ell`, la
sonnerie ASCII). L'API l'accepte tel quel, et le premier export XLSX - le
format PAR DEFAUT de `/api/crm/export/all` - rend un 500 :
`openpyxl.utils.exceptions.IllegalCharacterError: bell cannot be used in
worksheets`. Aucun fichier n'est produit, et les autres contacts, projets et
livrables partent avec.

Le format XML d'un classeur n'accepte pas ces octets ; ils n'ont non plus
aucun sens dans une fiche client. On les neutralise a l'endroit ou une valeur
devient une cellule, en gardant tabulation, retour a la ligne et retour
chariot, qui sont legitimes dans une note.
"""

from __future__ import annotations

import io

import pytest


async def _contact_empoisonne(client) -> None:
    reponse = await client.post(
        "/api/memory/contacts",
        json={
            "first_name": "Sonne",
            "last_name": "Rie",
            "company": "b\x07ell",
            "email": "sonnerie@test.fr",
        },
    )
    assert reponse.status_code == 200, reponse.text


class TestUnCaractereDeControleNEmporteRienDuTout:

    @pytest.mark.asyncio
    async def test_l_export_xlsx_aboutit_malgre_la_fiche_empoisonnee(self, client):
        await _contact_empoisonne(client)

        reponse = await client.post("/api/crm/export/all")

        assert reponse.status_code == 200, (
            f"une fiche empoisonnee ne doit pas emporter l'export entier : "
            f"{reponse.status_code} {reponse.text[:200]}"
        )
        assert len(reponse.content) > 0, "aucun fichier produit"

    @pytest.mark.asyncio
    async def test_la_valeur_est_neutralisee_pas_perdue(self, client):
        from openpyxl import load_workbook

        await _contact_empoisonne(client)

        reponse = await client.post("/api/crm/export/all")
        assert reponse.status_code == 200, reponse.text[:200]

        classeur = load_workbook(io.BytesIO(reponse.content))
        feuille = classeur["Contacts"]
        entetes = [cellule.value for cellule in feuille[1]]
        colonne = entetes.index("Entreprise")
        valeurs = [ligne[colonne] for ligne in feuille.iter_rows(min_row=2, values_only=True)]
        assert "bell" in valeurs, (
            f"la societe doit rester lisible, debarrassee du caractere de "
            f"controle : {valeurs}"
        )

    @pytest.mark.asyncio
    async def test_l_export_contacts_seul_aboutit_aussi(self, client):
        await _contact_empoisonne(client)

        reponse = await client.post("/api/crm/export/contacts?format=xlsx")

        assert reponse.status_code == 200, reponse.text[:200]

    @pytest.mark.asyncio
    async def test_les_sauts_de_ligne_d_une_note_sont_conserves(self, client):
        """La neutralisation ne doit pas raboter ce qui est legitime."""
        from app.services.crm_export import _format_value

        assert _format_value("ligne 1\nligne 2\tsuite\r") == "ligne 1\nligne 2\tsuite\r"
        assert _format_value("b\x07ell") == "bell"
        assert _format_value(None) == ""


class TestB449UneFormuleNeSortJamaisActiveDeLExport:
    """B-449 (05/09/2026) : une société « =1+1 » sortait de l'export XLSX en
    formule ACTIVE (data_type 'f') et du CSV telle quelle. La neutralisation
    à l'écriture est contournée par au moins deux constructeurs Contact ;
    l'export doit être le dernier rempart, sans abîmer un téléphone en +33.
    """

    async def _contact_formule(self, client) -> None:
        reponse = await client.post(
            "/api/memory/contacts",
            json={
                "first_name": "For",
                "last_name": "Mule",
                "company": "=1+1",
                "phone": "+33 6 12 34 56 78",
                "email": "formule@test.fr",
                "notes": '=HYPERLINK("http://evil.test","clique")',
            },
        )
        assert reponse.status_code == 200, reponse.text

    @pytest.mark.asyncio
    async def test_le_xlsx_stocke_une_chaine_pas_une_formule(self, client):
        from openpyxl import load_workbook

        await self._contact_formule(client)
        reponse = await client.post("/api/crm/export/contacts?format=xlsx")
        assert reponse.status_code == 200, reponse.text[:200]

        feuille = load_workbook(io.BytesIO(reponse.content))["Contacts"]
        entetes = [c.value for c in feuille[1]]
        col_societe = entetes.index("Entreprise") + 1
        col_tel = entetes.index("Telephone") + 1
        cellule = feuille.cell(row=2, column=col_societe)
        assert cellule.data_type != "f", "la formule est ACTIVE dans le classeur"
        assert "1+1" in str(cellule.value), "la valeur doit rester lisible"
        assert feuille.cell(row=2, column=col_tel).value == "+33 6 12 34 56 78", (
            "un téléphone en +33 ne doit pas être abîmé"
        )

    @pytest.mark.asyncio
    async def test_le_csv_neutralise_le_prefixe_de_formule(self, client):
        await self._contact_formule(client)
        reponse = await client.post("/api/crm/export/contacts?format=csv")
        assert reponse.status_code == 200, reponse.text[:200]
        texte = reponse.content.decode("utf-8-sig")
        ligne = next(candidate for candidate in texte.splitlines() if "formule@test.fr" in candidate)
        assert "=1+1" not in ligne.replace("'=1+1", ""), ligne
        assert "'=1+1" in ligne or '"\'=1+1"' in ligne, f"préfixe de formule non neutralisé : {ligne}"
        assert "+33 6 12 34 56 78" in ligne
